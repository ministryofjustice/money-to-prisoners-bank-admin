from collections import defaultdict

from openpyxl import load_workbook, styles
from openpyxl.writer.excel import save_virtual_workbook


ADI_JOURNAL_LABEL = 'ADI_JOURNAL'
ACCESSPAY_LABEL = 'ACCESSPAY_REFUNDS'
MT940_STMT_LABEL = 'MT940_BANK_STMT'


class Journal():

    STYLE_TYPES = {
        'fill': styles.PatternFill,
        'border': styles.Border,
        'font': styles.Font,
        'alignment': styles.Alignment
    }

    def __init__(self, template_path, sheet_name, start_row, fields):
        self.wb = load_workbook(template_path, keep_vba=True)
        self.journal_ws = self.wb.get_sheet_by_name(sheet_name)

        self.start_row = start_row
        self.current_row = start_row
        self.fields = fields

    def next_row(self, increment=1):
        self.current_row += increment

    def get_cell(self, field):
        return '%s%s' % (self.fields[field]['column'],
                         self.current_row)

    def set_field(self, field, value, style=None, extra_style={}):
        cell = self.get_cell(field)
        self.journal_ws[cell] = value

        computed_style = defaultdict(dict)
        base_style = style or self.fields[field].get('style', {})
        for key in base_style:
            computed_style[key].update(base_style[key])

        for key in extra_style:
            computed_style[key].update(extra_style[key])

        for key in computed_style:
            setattr(
                self.journal_ws[cell],
                key,
                self.STYLE_TYPES[key](**computed_style[key])
            )
        return self.journal_ws[cell]

    def lookup(self, field, context={}):
        try:
            value = self.fields[field]['value']
            return value.format(**context)
        except KeyError:
            pass  # no static value
        return None

    def create_file(self):
        return save_virtual_workbook(self.wb)
