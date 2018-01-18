from collections import defaultdict
from datetime import date, datetime, timedelta
import logging
from unittest import mock, skip
from urllib.parse import quote_plus

from django.test.client import RequestFactory
from django.core.urlresolvers import reverse
from django.utils.timezone import utc
from mtp_common.auth.api_client import get_api_session
from mtp_common.auth.models import MojUser
from mtp_common.test_utils import silence_logger
from openpyxl import load_workbook
import responses

from .utils import (
    TEST_PRISONS, NO_TRANSACTIONS, mock_list_prisons,
    get_test_transactions, get_test_credits, temp_file, api_url,
    mock_bank_holidays, ResponsesTestCase
)
from bank_admin import adi, adi_config
from bank_admin.exceptions import EmptyFileError, EarlyReconciliationError
from bank_admin.types import PaymentType
from bank_admin.utils import set_worldpay_cutoff


def get_cell_value(journal_ws, field, row):
    cell = '%s%s' % (
        adi_config.ADI_JOURNAL_FIELDS[field]['column'],
        row
    )
    return journal_ws[cell].value


class AdiPaymentFileGenerationTestCase(ResponsesTestCase):

    def setUp(self):
        self.factory = RequestFactory()

    def get_api_session(self, **kwargs):
        request = self.factory.get(
            reverse('bank_admin:download_adi_journal'),
            **kwargs
        )
        request.user = MojUser(
            1, '',
            {'first_name': 'John', 'last_name': 'Smith', 'username': 'jsmith'}
        )
        request.session = mock.MagicMock()
        return get_api_session(request)

    def get_user(self):
        return MojUser(
            1, '',
            {'first_name': 'John', 'last_name': 'Smith', 'username': 'jsmith'}
        )

    def _generate_test_adi_journal(self, receipt_date=None, user=None):
        if receipt_date is None:
            receipt_date = date(2016, 9, 13)
        start_date = quote_plus(str(set_worldpay_cutoff(receipt_date)))
        end_date = quote_plus(str(set_worldpay_cutoff(receipt_date + timedelta(days=1))))

        credits = get_test_credits(20)
        refundable_transactions = get_test_transactions(PaymentType.refund, 5)
        rejected_transactions = get_test_transactions(PaymentType.reject, 2)

        responses.add(
            responses.POST,
            api_url('/transactions/reconcile/'),
            status=200
        )
        mock_list_prisons()
        responses.add(
            responses.GET,
            api_url('/credits/'),
            json=credits
        )
        responses.add(
            responses.GET,
            api_url(
                '/transactions/?offset=0&limit=500'
                '&received_at__lt={end_date}'
                '&received_at__gte={start_date}'
                '&status=refundable'.format(
                    start_date=start_date, end_date=end_date
                )
            ),
            json=refundable_transactions,
            match_querystring=True
        )
        responses.add(
            responses.GET,
            api_url(
                '/transactions/?offset=0&limit=500'
                '&received_at__lt={end_date}'
                '&received_at__gte={start_date}'
                '&status=unidentified'.format(
                    start_date=start_date, end_date=end_date
                )
            ),
            json=rejected_transactions,
            match_querystring=True
        )
        mock_bank_holidays()

        with silence_logger(name='mtp', level=logging.WARNING):
            filename, exceldata = adi.generate_adi_journal(
                self.get_api_session(), receipt_date, user=user
            )

        return filename, exceldata, (credits, refundable_transactions, rejected_transactions)

    def _get_expected_number_of_rows(self, credits, refundable_transactions, rejected_transactions):
        expected_credits = 0
        card_payment_ref_codes = set()
        for credit in credits['results']:
            if credit['source'] == 'bank_transfer':
                expected_credits += 1
            else:
                # one lump sum for each card payment batch
                if credit['reconciliation_code'] not in card_payment_ref_codes:
                    card_payment_ref_codes.add(credit['reconciliation_code'])
                    expected_credits += 1

        expected_debit_rows = (
            # valid credits
            expected_credits +
            # refunds
            len(refundable_transactions['results']) +
            # rejects
            len(rejected_transactions['results'])
        )

        expected_credit_rows = (
            # valid credits
            len({prison['general_ledger_code'] for prison in TEST_PRISONS}) +
            # refunds
            1 +
            # rejects
            len(rejected_transactions['results'])
        )

        return expected_debit_rows, expected_credit_rows

    @responses.activate
    @skip('Enable to generate an example file for inspection')
    def test_adi_journal_generation(self):
        filename, exceldata, _ = self._generate_test_adi_journal()
        with open(filename, 'wb+') as f:
            f.write(exceldata)

    @responses.activate
    def test_adi_journal_debits_match_credits(self):
        filename, exceldata, test_data = self._generate_test_adi_journal()
        credits, refundable_transactions, rejected_transactions = test_data

        with temp_file(exceldata) as f:
            wb = load_workbook(f)
            journal_ws = wb.get_sheet_by_name('130916')
            row = adi_config.ADI_JOURNAL_START_ROW

            current_balance = 0
            while True:
                debit = get_cell_value(journal_ws, 'debit', row)
                credit = get_cell_value(journal_ws, 'credit', row)

                if debit and credit:
                    # final line
                    break
                elif debit:
                    current_balance -= debit
                elif credit:
                    current_balance += credit
                row += 1
            self.assertAlmostEqual(0, current_balance)

    @responses.activate
    def test_adi_journal_number_of_payment_rows_correct(self):
        filename, exceldata, test_data = self._generate_test_adi_journal()
        credits, refundable_transactions, rejected_transactions = test_data

        with temp_file(exceldata) as f:
            wb = load_workbook(f)
            journal_ws = wb.get_sheet_by_name('130916')
            row = adi_config.ADI_JOURNAL_START_ROW

            expected_debit_rows, expected_credit_rows = self._get_expected_number_of_rows(
                credits, refundable_transactions, rejected_transactions
            )

            file_debit_rows = 0
            file_credit_rows = 0
            while True:
                debit = get_cell_value(journal_ws, 'debit', row)
                credit = get_cell_value(journal_ws, 'credit', row)

                if debit and credit:
                    # final line
                    break
                elif debit:
                    file_debit_rows += 1
                elif credit:
                    file_credit_rows += 1
                row += 1

            credit = self.assertEqual(expected_credit_rows, file_credit_rows)
            credit = self.assertEqual(expected_debit_rows, file_debit_rows)

    @responses.activate
    def test_adi_journal_credit_sums_correct(self):
        filename, exceldata, test_data = self._generate_test_adi_journal()
        credits, refundable_transactions, rejected_transactions = test_data

        prison_totals = defaultdict(int)
        for prison in TEST_PRISONS:
            prison_totals[prison['general_ledger_code']] += float(sum(
                [c['amount'] for c in credits['results']
                    if 'prison' in c and c['prison'] == prison['nomis_id']]
            ))/100

        refund_total = float(sum([t['amount'] for t in refundable_transactions['results']]))/100

        with temp_file(exceldata) as f:
            wb = load_workbook(f)
            journal_ws = wb.get_sheet_by_name('130916')
            row = adi_config.ADI_JOURNAL_START_ROW

            refund_bu_code = adi_config.ADI_JOURNAL_FIELDS['cost_centre']['value']['refund']['credit']
            reject_bu_code = adi_config.ADI_JOURNAL_FIELDS['cost_centre']['value']['reject']['credit']

            while True:
                debit = get_cell_value(journal_ws, 'debit', row)
                credit = get_cell_value(journal_ws, 'credit', row)

                if debit and credit:
                    # final line
                    break
                elif credit:
                    bu_code = get_cell_value(journal_ws, 'cost_centre', row)
                    if bu_code == refund_bu_code:
                        self.assertAlmostEqual(credit, refund_total)
                    elif bu_code != reject_bu_code:
                        self.assertAlmostEqual(credit, prison_totals[bu_code])
                row += 1

    @responses.activate
    def test_no_transactions_raises_error(self):
        responses.add(
            responses.POST,
            api_url('/transactions/reconcile/'),
            status=200
        )
        mock_list_prisons()
        responses.add(
            responses.GET,
            api_url('/credits/'),
            json=NO_TRANSACTIONS
        )
        responses.add(
            responses.GET,
            api_url('/transactions/'),
            json=NO_TRANSACTIONS
        )
        responses.add(
            responses.GET,
            api_url('/transactions/'),
            json=NO_TRANSACTIONS
        )
        mock_bank_holidays()

        with self.assertRaises(EmptyFileError), silence_logger(name='mtp', level=logging.WARNING):
            _, exceldata = adi.generate_adi_journal(self.get_api_session(), date(2016, 9, 13))

    @responses.activate
    def test_adi_journal_reconciles_date(self):
        _, _, _ = self._generate_test_adi_journal(receipt_date=date(2016, 9, 13))

        self.assert_called_with(
            api_url('/transactions/reconcile/'), responses.POST,
            {
                'received_at__gte': datetime(2016, 9, 13, 0, 0, tzinfo=utc).isoformat(),
                'received_at__lt': datetime(2016, 9, 14, 0, 0, tzinfo=utc).isoformat()
            }
        )

    @responses.activate
    def test_adi_journal_upload_range_set(self):
        filename, exceldata, test_data = self._generate_test_adi_journal()
        credits, refundable_transactions, rejected_transactions = test_data

        expected_debit_rows, expected_credit_rows = self._get_expected_number_of_rows(
            credits, refundable_transactions, rejected_transactions
        )
        expected_rows = expected_debit_rows + expected_credit_rows

        with temp_file(exceldata) as f:
            wb = load_workbook(f)
            journal_ws = wb.get_sheet_by_name('130916')

            self.assertEqual(
                list(wb.get_named_range('BNE_UPLOAD').destinations),
                [(
                    journal_ws.title,
                    '$B$%(start)s:$B$%(end)s' % {
                        'start': adi_config.ADI_JOURNAL_START_ROW,
                        'end': adi_config.ADI_JOURNAL_START_ROW + expected_rows - 1
                    }
                )]
            )

    @responses.activate
    def test_batch_name_includes_initials(self):
        filename, exceldata, _ = self._generate_test_adi_journal(user=self.get_user())

        with temp_file(exceldata) as f:
            wb = load_workbook(f)
            journal_ws = wb.get_sheet_by_name('130916')
            self.assertTrue('JS' in journal_ws[adi_config.ADI_BATCH_NAME_CELL].value)

    @responses.activate
    @mock.patch('bank_admin.adi.date')
    def test_accounting_date_is_download_date(self, mock_date):
        processing_date = date(2016, 8, 31)
        mock_date.today.return_value = processing_date
        receipt_date = date(2016, 8, 30)
        filename, exceldata, _ = self._generate_test_adi_journal(
            receipt_date=receipt_date
        )

        with temp_file(exceldata) as f:
            wb = load_workbook(f)
            journal_ws = wb.get_sheet_by_name(receipt_date.strftime('%d%m%y'))
            self.assertEqual(
                processing_date.strftime(adi_config.ADI_DATE_FORMAT),
                journal_ws[adi_config.ADI_DATE_CELL].value
            )

    @responses.activate
    @mock.patch('bank_admin.adi.date')
    def test_accounting_date_is_set_back_across_month_boundary(self, mock_date):
        processing_date = date(2016, 9, 1)
        mock_date.today.return_value = processing_date
        receipt_date = date(2016, 8, 31)
        filename, exceldata, _ = self._generate_test_adi_journal(
            receipt_date=receipt_date
        )

        with temp_file(exceldata) as f:
            wb = load_workbook(f)
            journal_ws = wb.get_sheet_by_name(receipt_date.strftime('%d%m%y'))
            self.assertEqual(
                receipt_date.strftime(adi_config.ADI_DATE_FORMAT),
                journal_ws[adi_config.ADI_DATE_CELL].value
            )

    @responses.activate
    def test_early_reconciliation_raises_error(self):
        with self.assertRaises(EarlyReconciliationError):
            self._generate_test_adi_journal(receipt_date=date.today())
