from collections import defaultdict
from datetime import datetime, time, timedelta, timezone
import io
from itertools import count, islice
import time as systime
import os

from django.utils.timezone import now
from mtp_common.api import retrieve_all_pages_for_path
from mtp_common.dates import WorkdayChecker
from openpyxl import load_workbook, styles
from openpyxl.writer.excel import save_workbook

from .exceptions import EarlyReconciliationError


def retrieve_all_transactions(api_session, **kwargs):
    return retrieve_all_pages_for_path(
        api_session, 'transactions/', **kwargs)


def retrieve_all_valid_credits(api_session, **kwargs):
    return retrieve_all_pages_for_path(
        api_session, 'credits/', valid=True, **kwargs)


def retrieve_prisons(api_session):
    prisons = retrieve_all_pages_for_path(api_session, 'prisons/')
    return {prison['nomis_id']: prison for prison in prisons}


def set_worldpay_cutoff(date):
    return datetime.combine(date, time(0, 0, 0, tzinfo=timezone.utc))


def get_start_and_end_date(date):
    checker = WorkdayChecker()
    start_date = set_worldpay_cutoff(date)
    end_date = set_worldpay_cutoff(checker.get_next_workday(date))
    return start_date, end_date


def reconcile_for_date(api_session, receipt_date):
    start_date, end_date = get_start_and_end_date(receipt_date)

    if start_date.date() >= now().date() or end_date.date() > now().date():
        raise EarlyReconciliationError

    reconciliation_date = start_date
    while reconciliation_date < end_date:
        end_of_day = reconciliation_date + timedelta(days=1)
        api_session.post(
            'transactions/reconcile/',
            json={
                'received_at__gte': reconciliation_date.isoformat(),
                'received_at__lt': end_of_day.isoformat(),
            }
        )
        reconciliation_date = end_of_day

    return start_date, end_date


def retrieve_last_balance(api_session, date):
    response = api_session.get(
        'balances/', params={
            'limit': 1,
            'date__lt': date.isoformat()
        }
    ).json()
    if response.get('results'):
        return response['results'][0]
    else:
        return None


def get_daily_file_uid():
    return int(systime.time()) % 86400


def escape_csv_formula(value):
    """
    Escapes formulae (strings that start with =) to prevent
    spreadsheet software vulnerabilities being exploited
    :param value: the value being added to a CSV cell
    """
    if isinstance(value, str) and value.startswith('='):
        return "'" + value
    return value


def get_full_narrative(transaction):
    return ' '.join([
        str(transaction[field_name]) for field_name
        in ['sender_name', 'reference']
        if transaction.get(field_name)
    ])


def get_preceding_workday_list(number_of_days, offset=0):
    """
    Returns a list of weekdays counting backwards from today
    :param number_of_days: number of weekdays to include in total
    :param offset: number of days ago to start from; if 0 today is included
    """

    def day_generator():
        current = now().date()
        for day in count():
            yield current - timedelta(days=day)

    days = day_generator()
    checker = WorkdayChecker()
    days = filter(checker.is_workday, days)
    days = islice(days, offset, number_of_days + offset)

    return list(days)


class Journal:

    STYLE_TYPES = {
        'fill': styles.PatternFill,
        'border': styles.Border,
        'font': styles.Font,
        'alignment': styles.Alignment
    }

    def __init__(self, template_path, sheet_name, start_row, fields):
        self.wb = load_workbook(template_path, keep_vba=True)
        self.journal_ws = self.wb[sheet_name]

        self.start_row = start_row
        self.current_row = start_row
        self.fields = fields

    def next_row(self, increment=1):
        self.current_row += increment

    def get_cell(self, field):
        return '%s%s' % (self.fields[field]['column'],
                         self.current_row)

    def set_field(self, field, value, style=None, extra_style=None):
        extra_style = extra_style or {}
        cell = self.get_cell(field)
        self.journal_ws[cell] = value

        computed_style = defaultdict(dict)
        base_style = style or self.fields[field].get('style', {})
        for key in base_style:
            computed_style[key].update(base_style[key])

        for key in extra_style:
            computed_style[key].update(extra_style[key])

        for key in computed_style:
            setattr(
                self.journal_ws[cell],
                key,
                self.STYLE_TYPES[key](**computed_style[key])
            )
        return self.journal_ws[cell]

    def lookup(self, field, context=None):
        context = context or {}

        try:
            value = self.fields[field]['value']
            return value.format(**context)
        except KeyError:
            pass  # no static value
        return None

    def create_file(self):
        f = io.BytesIO()
        save_workbook(self.wb, f)
        return f.getvalue()


def get_cached_file_path(label, date, extension=None):
    filepath = 'local_files/cache/{label}/{date:%Y%m%d}'.format(label=label, date=date)
    if extension:
        filepath = '.'.join([filepath, extension])
    return filepath


def get_or_create_file(label, date, creation_func, f_args=None, f_kwargs=None, file_extension=None):
    f_args = f_args or []
    f_kwargs = f_kwargs or {}

    filepath = get_cached_file_path(label, date, extension=file_extension)
    if not os.path.isfile(filepath):
        filedata = creation_func(*f_args, **f_kwargs)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'wb+') as f:
            if isinstance(filedata, str):
                filedata = filedata.encode('utf-8')
            f.write(filedata)
    return filepath
